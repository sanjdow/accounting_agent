"""Generate the close package: PDF summary + Excel workbook with raw data."""
from __future__ import annotations
from pathlib import Path
from typing import Any, Dict

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
)
from reportlab.lib import colors

from config import REPORTS_DIR


def _fmt(v) -> str:
    if isinstance(v, float):
        return f"{v:,.2f}"
    return str(v)


def generate_pdf(run_id: str, snapshot: Dict[str, Any]) -> Path:
    out = REPORTS_DIR / f"close_{run_id}.pdf"
    doc = SimpleDocTemplate(
        str(out), pagesize=A4,
        leftMargin=18*mm, rightMargin=18*mm, topMargin=18*mm, bottomMargin=18*mm,
    )
    styles = getSampleStyleSheet()
    h1 = styles["Heading1"]; h2 = styles["Heading2"]; body = styles["BodyText"]
    mono = ParagraphStyle("mono", parent=body, fontName="Courier", fontSize=8, leading=10)

    flow = []
    flow.append(Paragraph(f"Month-End Close Package — run {run_id}", h1))
    flow.append(Paragraph(f"Period: <b>{snapshot.get('period','')}</b>", body))
    flow.append(Spacer(1, 6))

    # --- Totals
    tb = snapshot["snapshot"]["tb_totals"]
    gl = snapshot["snapshot"]["gl_totals"]
    rec = snapshot["snapshot"]["reconciliation"]
    flow.append(Paragraph("Totals & Reconciliation", h2))
    t = Table([
        ["", "Debit", "Credit", "Diff"],
        ["TB", _fmt(tb["debit"]), _fmt(tb["credit"]), _fmt(tb["diff"])],
        ["GL", _fmt(gl["debit"]), _fmt(gl["credit"]), ""],
        ["TB - GL", _fmt(rec["tb_minus_gl_debit"]), _fmt(rec["tb_minus_gl_credit"]),
         "MATCHED" if rec["matched"] else "DIFF"],
    ], hAlign="LEFT")
    t.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
        ("GRID", (0,0), (-1,-1), 0.25, colors.grey),
        ("FONTSIZE", (0,0), (-1,-1), 9),
    ]))
    flow.append(t); flow.append(Spacer(1, 10))

    # --- Top accounts
    flow.append(Paragraph("Top Accounts by Net Movement", h2))
    rows = [["Account", "Name", "Debit", "Credit", "Net"]]
    for r in snapshot["snapshot"].get("top_accounts", [])[:10]:
        rows.append([r.get("account",""), r.get("account_name",""),
                     _fmt(r.get("debit",0)), _fmt(r.get("credit",0)), _fmt(r.get("net",0))])
    if len(rows) > 1:
        t = Table(rows, hAlign="LEFT")
        t.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),colors.lightgrey),
                               ("GRID",(0,0),(-1,-1),0.25,colors.grey),
                               ("FONTSIZE",(0,0),(-1,-1),8)]))
        flow.append(t)
    flow.append(Spacer(1, 10))

    # --- Anomalies
    flow.append(Paragraph("Anomalies", h2))
    anoms = snapshot.get("anomalies", [])
    if anoms:
        rows = [["Type", "Account", "Severity", "Detail"]]
        for a in anoms[:20]:
            detail = (f"z={a.get('z_score')}" if a.get("z_score") is not None
                      else f"score={a.get('anomaly_score')}")
            rows.append([a.get("type",""), a.get("account",""),
                         a.get("severity",""), detail])
        t = Table(rows, hAlign="LEFT")
        t.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),colors.lightgrey),
                               ("GRID",(0,0),(-1,-1),0.25,colors.grey),
                               ("FONTSIZE",(0,0),(-1,-1),8)]))
        flow.append(t)
    else:
        flow.append(Paragraph("No anomalies flagged.", body))
    flow.append(Spacer(1, 10))

    # --- Narrative
    flow.append(PageBreak())
    flow.append(Paragraph("Narrative", h2))
    for para in (snapshot.get("narrative","") or "").split("\n"):
        if para.strip():
            flow.append(Paragraph(para.replace("<","&lt;").replace(">","&gt;"), body))
            flow.append(Spacer(1, 4))

    # --- Policy references
    flow.append(Spacer(1, 8))
    flow.append(Paragraph("Policy References", h2))
    for h in snapshot.get("policy_hits", [])[:4]:
        flow.append(Paragraph(
            f"<b>{h['source']}</b> (confidence {h.get('confidence',0)})", body
        ))
        flow.append(Paragraph(h["text"][:400].replace("<","&lt;").replace(">","&gt;"), mono))
        flow.append(Spacer(1, 4))

    doc.build(flow)
    return out


def generate_xlsx(run_id: str, snapshot: Dict[str, Any],
                  tb_df: pd.DataFrame, gl_df: pd.DataFrame) -> Path:
    out = REPORTS_DIR / f"close_{run_id}.xlsx"
    wb = Workbook()
    ws = wb.active; ws.title = "Summary"

    s = snapshot["snapshot"]
    ws["A1"] = "Month-End Close"; ws["B1"] = run_id
    ws["A2"] = "Period"; ws["B2"] = s.get("period","")
    ws["A4"] = "TB Debit"; ws["B4"] = s["tb_totals"]["debit"]
    ws["A5"] = "TB Credit"; ws["B5"] = s["tb_totals"]["credit"]
    ws["A6"] = "TB Diff"; ws["B6"] = s["tb_totals"]["diff"]
    ws["A7"] = "GL Debit"; ws["B7"] = s["gl_totals"]["debit"]
    ws["A8"] = "GL Credit"; ws["B8"] = s["gl_totals"]["credit"]
    ws["A9"] = "Reconciled"; ws["B9"] = "YES" if s["reconciliation"]["matched"] else "NO"

    def _add_sheet(name: str, df: pd.DataFrame):
        sh = wb.create_sheet(name)
        if df.empty:
            sh["A1"] = "(no data)"
            return
        for row in dataframe_to_rows(df, index=False, header=True):
            sh.append(row)

    _add_sheet("TB", tb_df)
    _add_sheet("GL", gl_df)
    _add_sheet("Top_Accounts", pd.DataFrame(s.get("top_accounts", [])))
    _add_sheet("Anomalies", pd.DataFrame(snapshot.get("anomalies", [])))
    _add_sheet("Account_Balances", pd.DataFrame(s.get("account_balances", [])))
    # Forecasts as a flat table
    fc = snapshot.get("forecasts", {})
    fc_rows = [{"key": k, **(v if isinstance(v, dict) else {"value": v})}
               for k, v in fc.items()]
    _add_sheet("Forecasts", pd.DataFrame(fc_rows))
    _add_sheet("Policy_Hits", pd.DataFrame(snapshot.get("policy_hits", [])))

    wb.save(str(out))
    return out
